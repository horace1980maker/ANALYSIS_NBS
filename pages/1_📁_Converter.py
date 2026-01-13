"""
NbS Analyzer — Page 1: Data Converter

Transform raw Excel files into analysis-ready format.
"""

from __future__ import annotations

import io
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


# =============================================================================
# Page Configuration
# =============================================================================

st.set_page_config(
    page_title="NbS Analyzer — Converter",
    page_icon="📁",
    layout="wide",
)

# Shared CSS (imported from app.py style)
st.markdown("""
<style>
    :root {
        --primary: #10b981;
        --primary-dark: #059669;
        --secondary: #6366f1;
        --bg-card: rgba(30, 41, 59, 0.7);
        --text-primary: #f1f5f9;
        --text-secondary: #94a3b8;
        --border: rgba(148, 163, 184, 0.2);
        --glass: rgba(255, 255, 255, 0.05);
        --success: #22c55e;
        --error: #ef4444;
    }
    
    .stApp {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 50%, #0f172a 100%);
    }
    
    #MainMenu, footer {visibility: hidden;}
    
    .glass-card {
        background: var(--bg-card);
        backdrop-filter: blur(20px);
        border: 1px solid var(--border);
        border-radius: 16px;
        padding: 2rem;
        margin: 1rem 0;
    }
    
    .hero {
        text-align: center;
        padding: 2rem;
        margin-bottom: 2rem;
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.1), rgba(99, 102, 241, 0.1));
        border-radius: 20px;
        border: 1px solid var(--border);
    }
    
    .hero h1 {
        font-size: 2rem;
        background: linear-gradient(135deg, #10b981, #6366f1);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    
    .hero p { color: var(--text-secondary); }
    
    .stats-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(120px, 1fr));
        gap: 1rem;
        margin: 1.5rem 0;
    }
    
    .stat-card {
        background: var(--glass);
        border: 1px solid var(--border);
        border-radius: 12px;
        padding: 1rem;
        text-align: center;
    }
    
    .stat-value { font-size: 1.5rem; font-weight: 700; color: var(--primary); }
    .stat-label { font-size: 0.75rem; color: var(--text-secondary); }
    
    .status-success {
        background: rgba(34, 197, 94, 0.2);
        border: 1px solid var(--success);
        color: var(--success);
        padding: 1rem;
        border-radius: 12px;
        margin: 1rem 0;
    }
    
    .status-error {
        background: rgba(239, 68, 68, 0.2);
        border: 1px solid var(--error);
        color: var(--error);
        padding: 1rem;
        border-radius: 12px;
        margin: 1rem 0;
    }
    
    .info-box {
        background: rgba(99, 102, 241, 0.1);
        border-left: 4px solid var(--secondary);
        padding: 1rem 1.5rem;
        border-radius: 0 12px 12px 0;
        margin: 1rem 0;
    }
    
    .info-box h4 { color: #a5b4fc; margin-bottom: 0.5rem; }
    .info-box p, .info-box li { color: var(--text-secondary); font-size: 0.9rem; }
    
    .stButton > button {
        background: linear-gradient(135deg, var(--primary), var(--primary-dark));
        color: white;
        border: none;
        border-radius: 12px;
        padding: 0.75rem 2rem;
        font-weight: 600;
    }
    
    .stDownloadButton > button {
        background: linear-gradient(135deg, var(--secondary), #4f46e5);
        color: white;
        border: none;
        border-radius: 12px;
        width: 100%;
    }
    
    .step-indicator {
        display: flex;
        justify-content: center;
        gap: 2rem;
        margin: 1.5rem 0;
    }
    
    .step {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        padding: 0.75rem 1.25rem;
        background: var(--glass);
        border-radius: 10px;
        border: 1px solid var(--border);
    }
    
    .step.active { background: rgba(16, 185, 129, 0.2); border-color: var(--primary); }
    .step.completed { background: rgba(34, 197, 94, 0.2); border-color: var(--success); }
    
    .step-number {
        width: 28px; height: 28px;
        border-radius: 50%;
        background: var(--border);
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 600;
        font-size: 0.85rem;
        color: var(--text-primary);
    }
    
    .step.active .step-number { background: var(--primary); }
    .step.completed .step-number { background: var(--success); }
    .step-text { color: var(--text-secondary); font-size: 0.85rem; }
    .step.active .step-text, .step.completed .step-text { color: var(--text-primary); }
</style>
""", unsafe_allow_html=True)


# =============================================================================
# Helper Functions
# =============================================================================

def remove_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def clean_col(col: str) -> str:
    col = (col or "").strip()
    col = remove_accents(col)
    col = col.replace("–", "-")
    col = re.sub(r"[^0-9A-Za-z]+", "_", col)
    col = re.sub(r"_+", "_", col).strip("_")
    col = col.lower()
    if re.match(r"^\d", col):
        col = "x" + col
    return col


def make_unique(values: List[str]) -> List[str]:
    seen = {}
    out = []
    for v in values:
        if v not in seen:
            seen[v] = 1
            out.append(v)
        else:
            seen[v] += 1
            out.append(f"{v}_{seen[v]}")
    return out


THREAT_CODE_RE = re.compile(r"\b(AC\d{2}|ANC\d{2})\b")
GAP_CODE_RE = re.compile(r"\bB?(\d+\.\d+)\b")


def extract_threat_code(text: Any) -> Optional[str]:
    if not isinstance(text, str):
        return None
    m = THREAT_CODE_RE.search(text)
    return m.group(1) if m else None


def extract_gap_code(text: Any) -> Optional[str]:
    if not isinstance(text, str):
        return None
    m = GAP_CODE_RE.search(text)
    return "B" + m.group(1) if m else None


def add_df_sheet(wb, title, df, table_style="TableStyleMedium9"):
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    
    header_fill = PatternFill("solid", fgColor="FFB7E1CD")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if len(df) > 0 else 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 60)
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def extract_lookup_threats(wb, sheet_name="AN - ANC"):
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet_name]
    rows = []
    for r in range(15, 200):
        code = ws.cell(r, 2).value
        if isinstance(code, str) and re.match(r"^AC\d{2}$", code.strip()):
            rows.append({"threat_code": code.strip(), "threat_type": "climatic", "group": ws.cell(r, 3).value, "threat": ws.cell(r, 4).value, "examples": ws.cell(r, 5).value})
        elif r > 25 and code is None:
            break
    for r in range(34, 400):
        code = ws.cell(r, 2).value
        if isinstance(code, str) and re.match(r"^ANC\d{2}$", code.strip()):
            rows.append({"threat_code": code.strip(), "threat_type": "non_climatic", "group": ws.cell(r, 3).value, "threat": ws.cell(r, 4).value, "examples": ws.cell(r, 5).value})
        elif r > 56 and ws.cell(r, 2).value is None:
            break
    return pd.DataFrame(rows)


def extract_lookup_security(wb, sheet_name="Seguridad"):
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet_name]
    rows = []
    for r in range(14, 30):
        var = ws.cell(r, 3).value
        if isinstance(var, str) and var.startswith("Seg_"):
            rows.append({"variable_raw": var, "variable": clean_col(var), "covers": ws.cell(r, 4).value})
    return pd.DataFrame(rows)


def extract_lookup_traits(wb, sheet_name="Rasgos Transformadores"):
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet_name]
    short_defs = {}
    for r in range(13, 19):
        trait = ws.cell(r, 2).value
        if isinstance(trait, str) and trait.startswith("T"):
            short_defs[clean_col(trait)] = {"trait_raw": trait, "short_definition": ws.cell(r, 3).value}
    long_defs = {}
    for r in range(26, 32):
        trait = ws.cell(r, 2).value
        if isinstance(trait, str) and trait.startswith("T"):
            long_defs[clean_col(trait)] = ws.cell(r, 3).value
    return pd.DataFrame([{"trait_raw": d["trait_raw"], "trait": t, "short_definition": d["short_definition"], "long_definition": long_defs.get(t)} for t, d in short_defs.items()])


def extract_lookup_gov_gaps(wb, sheet_name="Brechas Gobernanza"):
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet_name]
    rows = []
    current_group = None
    for r in range(1, 400):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.strip().lower().startswith("retos de"):
            current_group = v.strip()
        if isinstance(v, str) and re.match(r"^\d+\.\d+\s", v.strip()):
            code = v.strip().split(" ")[0]
            name = " ".join(v.strip().split(" ")[1:])
            rows.append({"gap_code": "B" + code, "gap_group": current_group, "gap_name": name, "gap_description": ws.cell(r, 3).value})
    df = pd.DataFrame(rows)
    return df.drop_duplicates(subset=["gap_code"]) if not df.empty else df


def transform_workbook(input_bytes: bytes, main_sheet: str) -> Tuple[bytes, Dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(input_bytes))
    df_raw = pd.read_excel(io.BytesIO(input_bytes), sheet_name=main_sheet)
    
    raw_cols = list(df_raw.columns)
    cleaned = make_unique([clean_col(c) for c in raw_cols])
    col_map = dict(zip(raw_cols, cleaned))
    fact = df_raw.rename(columns=col_map).copy()
    
    # Standardize missing
    missing_markers = {"ND", "N/D", "NA", "N.A.", "", " ", "nd", "n/d", "na"}
    for c in fact.columns:
        if fact[c].dtype == object:
            fact[c] = fact[c].apply(lambda x: np.nan if isinstance(x, str) and x.strip() in missing_markers else x)
            fact[c] = fact[c].apply(lambda x: x.strip() if isinstance(x, str) else x)
    
    # Numeric conversions
    for c in [c for c in fact.columns if c.endswith("_n")] + ["seg_total", "t_score"]:
        if c in fact.columns:
            fact[c] = pd.to_numeric(fact[c], errors="coerce")
    
    seg_cols = [c for c in fact.columns if c.startswith("seg_") and c not in ("seg_total", "seg_justificacion")]
    trait_cols = [c for c in fact.columns if re.match(r"^t[1-6]_", c)]
    for c in seg_cols + trait_cols:
        fact[c] = pd.to_numeric(fact[c], errors="coerce").astype("Int64")
    
    # Lookups
    lookup_threats = extract_lookup_threats(wb)
    lookup_security = extract_lookup_security(wb)
    lookup_traits = extract_lookup_traits(wb)
    lookup_gov = extract_lookup_gov_gaps(wb)
    
    threat_map = lookup_threats.set_index("threat_code")[["threat_type", "group", "threat"]].to_dict(orient="index") if not lookup_threats.empty else {}
    gap_map = lookup_gov.set_index("gap_code")[["gap_group", "gap_name"]].to_dict(orient="index") if not lookup_gov.empty else {}
    sec_map = lookup_security.set_index("variable")[["covers"]].to_dict(orient="index") if not lookup_security.empty else {}
    trait_map = lookup_traits.set_index("trait")[["short_definition"]].to_dict(orient="index") if not lookup_traits.empty else {}
    
    if "id_nbs" not in fact.columns:
        raise ValueError("Expected an 'ID_NbS' column in the main sheet")
    
    # TIDY_THREATS
    clim_cols = [c for c in fact.columns if c.startswith("amenaza_climatica_")]
    non_cols = [c for c in fact.columns if c.startswith("amenaza_no_climatica_")]
    threat_records = []
    for col in clim_cols + non_cols:
        fallback = "climatic" if col in clim_cols else "non_climatic"
        for _, row in fact[["id_nbs", col]].iterrows():
            val = row[col]
            if pd.isna(val):
                continue
            code = extract_threat_code(val)
            rec = {"id_nbs": row["id_nbs"], "source_column": col, "threat_text_raw": val, "threat_code": code, "threat_type": fallback}
            if code and code in threat_map:
                rec.update({"threat_type": threat_map[code]["threat_type"], "threat_group": threat_map[code]["group"], "threat_label": threat_map[code]["threat"]})
            else:
                rec.update({"threat_group": None, "threat_label": None})
            threat_records.append(rec)
    tidy_threats = pd.DataFrame(threat_records)
    
    # TIDY_GOV_GAPS
    gap_cols = [c for c in fact.columns if c.startswith("brecha_nbs_")]
    gap_records = []
    for col in gap_cols:
        for _, row in fact[["id_nbs", col]].iterrows():
            val = row[col]
            if pd.isna(val):
                continue
            code = extract_gap_code(val)
            rec = {"id_nbs": row["id_nbs"], "source_column": col, "gap_text_raw": val, "gap_code": code}
            if code and code in gap_map:
                rec.update({"gap_group": gap_map[code]["gap_group"], "gap_name": gap_map[code]["gap_name"]})
            else:
                rec.update({"gap_group": None, "gap_name": None})
            gap_records.append(rec)
    tidy_gov = pd.DataFrame(gap_records)
    
    # TIDY_SECURITY
    dim_vars = [v for v in (lookup_security["variable"].tolist() if not lookup_security.empty else []) if v not in ("seg_total", "seg_justificacion")]
    dim_cols = [c for c in fact.columns if c in dim_vars]
    tidy_security = fact[["id_nbs"] + dim_cols].melt(id_vars=["id_nbs"], var_name="dimension", value_name="value") if dim_cols else pd.DataFrame(columns=["id_nbs", "dimension", "value", "definition"])
    if not tidy_security.empty:
        tidy_security["value"] = pd.to_numeric(tidy_security["value"], errors="coerce")
        tidy_security = tidy_security.dropna(subset=["value"])
        tidy_security["definition"] = tidy_security["dimension"].map(lambda d: sec_map.get(d, {}).get("covers"))
    
    # TIDY_TRAITS
    trait_cols_clean = [c for c in fact.columns if c in (lookup_traits["trait"].tolist() if not lookup_traits.empty else [])]
    tidy_traits = fact[["id_nbs"] + trait_cols_clean].melt(id_vars=["id_nbs"], var_name="trait", value_name="value") if trait_cols_clean else pd.DataFrame(columns=["id_nbs", "trait", "value", "definition"])
    if not tidy_traits.empty:
        tidy_traits["value"] = pd.to_numeric(tidy_traits["value"], errors="coerce")
        tidy_traits = tidy_traits.dropna(subset=["value"])
        tidy_traits["definition"] = tidy_traits["trait"].map(lambda t: trait_map.get(t, {}).get("short_definition"))
    
    # QA
    missing_summary = pd.DataFrame({"column": fact.columns, "missing_n": [int(fact[c].isna().sum()) for c in fact.columns], "missing_pct": [float(fact[c].isna().mean()) for c in fact.columns]}).sort_values("missing_pct", ascending=False)
    column_map_df = pd.DataFrame({"raw_column": raw_cols, "clean_column": [col_map[c] for c in raw_cols]})
    
    # Output workbook
    out_wb = openpyxl.load_workbook(io.BytesIO(input_bytes))
    add_df_sheet(out_wb, "COLUMN_MAP", column_map_df)
    add_df_sheet(out_wb, "FACT_NBS", fact)
    add_df_sheet(out_wb, "TIDY_THREATS", tidy_threats)
    add_df_sheet(out_wb, "TIDY_GOV_GAPS", tidy_gov)
    add_df_sheet(out_wb, "TIDY_SECURITY", tidy_security)
    add_df_sheet(out_wb, "TIDY_TRAITS", tidy_traits)
    add_df_sheet(out_wb, "LOOKUP_THREATS", lookup_threats)
    add_df_sheet(out_wb, "LOOKUP_GOV_GAPS", lookup_gov)
    add_df_sheet(out_wb, "LOOKUP_SECURITY", lookup_security)
    add_df_sheet(out_wb, "LOOKUP_TRAITS", lookup_traits)
    add_df_sheet(out_wb, "QA_MISSINGNESS", missing_summary)
    
    output = io.BytesIO()
    out_wb.save(output)
    output.seek(0)
    
    stats = {
        "n_cases": len(fact),
        "n_columns": len(fact.columns),
        "n_threats": len(tidy_threats),
        "n_gaps": len(tidy_gov),
        "n_security": len(tidy_security),
        "n_traits": len(tidy_traits),
        "sheets_created": 11,
    }
    
    return output.getvalue(), stats


# =============================================================================
# UI
# =============================================================================

def render_step_indicator(current_step):
    steps = [("1", "Upload"), ("2", "Configure"), ("3", "Transform"), ("4", "Download")]
    html = '<div class="step-indicator">'
    for i, (num, text) in enumerate(steps, 1):
        status = "completed" if i < current_step else ("active" if i == current_step else "")
        html += f'<div class="step {status}"><div class="step-number">{num if i >= current_step else "✓"}</div><div class="step-text">{text}</div></div>'
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)


def render_stats(stats):
    st.markdown(f"""
    <div class="stats-grid">
        <div class="stat-card"><div class="stat-value">{stats['n_cases']}</div><div class="stat-label">NbS Cases</div></div>
        <div class="stat-card"><div class="stat-value">{stats['n_columns']}</div><div class="stat-label">Columns</div></div>
        <div class="stat-card"><div class="stat-value">{stats['n_threats']}</div><div class="stat-label">Threats</div></div>
        <div class="stat-card"><div class="stat-value">{stats['n_gaps']}</div><div class="stat-label">Gov Gaps</div></div>
        <div class="stat-card"><div class="stat-value">{stats['sheets_created']}</div><div class="stat-label">New Sheets</div></div>
    </div>
    """, unsafe_allow_html=True)


# Main
st.markdown("""
<div class="hero">
    <h1>📁 Data Converter</h1>
    <p>Transform your raw Excel file into an analysis-ready format</p>
</div>
""", unsafe_allow_html=True)

if "transformed_data" not in st.session_state:
    st.session_state.transformed_data = None
if "transform_stats" not in st.session_state:
    st.session_state.transform_stats = None
if "current_step" not in st.session_state:
    st.session_state.current_step = 1

render_step_indicator(st.session_state.current_step)

col1, col2, col3 = st.columns([1, 2, 1])

with col2:
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    
    st.markdown("### 📤 Upload Your Excel File")
    uploaded_file = st.file_uploader("Drag and drop or click to upload", type=["xlsx", "xls"], help="Upload your raw NbS analysis Excel file")
    
    if uploaded_file:
        st.session_state.current_step = 2
        try:
            xl = pd.ExcelFile(uploaded_file)
            sheets = xl.sheet_names
            
            st.markdown("### ⚙️ Configuration")
            default_sheet = next((s for s in sheets if "general" in s.lower() and "nbs" in s.lower()), sheets[0])
            main_sheet = st.selectbox("Select Main Data Sheet", sheets, index=sheets.index(default_sheet) if default_sheet in sheets else 0)
            
            st.markdown(f"""
            <div class="info-box">
                <h4>📋 File Info</h4>
                <p><strong>Filename:</strong> {uploaded_file.name}<br><strong>Sheets:</strong> {len(sheets)}<br><strong>Size:</strong> {uploaded_file.size / 1024:.1f} KB</p>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button("🔄 Transform Data", use_container_width=True):
                st.session_state.current_step = 3
                with st.spinner("Transforming..."):
                    try:
                        uploaded_file.seek(0)
                        transformed_bytes, stats = transform_workbook(uploaded_file.read(), main_sheet)
                        st.session_state.transformed_data = transformed_bytes
                        st.session_state.transform_stats = stats
                        st.session_state.current_step = 4
                        st.rerun()
                    except Exception as e:
                        st.markdown(f'<div class="status-error"><strong>❌ Error:</strong> {e}</div>', unsafe_allow_html=True)
        except Exception as e:
            st.error(f"Error reading file: {e}")
    
    if st.session_state.transformed_data:
        st.markdown('<div class="status-success"><strong>✓ Success!</strong> Transformation complete.</div>', unsafe_allow_html=True)
        
        if st.session_state.transform_stats:
            render_stats(st.session_state.transform_stats)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button("⬇️ Download Analysis-Ready File", st.session_state.transformed_data, f"NBS_analysis_ready_{timestamp}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        
        st.markdown("""
        <div class="info-box">
            <h4>📊 Next Step</h4>
            <p>Go to the <strong>Analyzer</strong> page to run Storyline A analysis on your converted file!</p>
        </div>
        """, unsafe_allow_html=True)
        
        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("🔄 Convert Another File"):
                st.session_state.transformed_data = None
                st.session_state.transform_stats = None
                st.session_state.current_step = 1
                st.rerun()
        with col_b:
            if st.button("📊 Go to Analyzer →"):
                st.switch_page("pages/2_📊_Analyzer.py")
    
    st.markdown('</div>', unsafe_allow_html=True)
